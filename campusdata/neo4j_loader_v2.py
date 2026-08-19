"""
neo4j_loader_v2.py
==================
jbnu_campus_graphdb_data.xlsx (마스터 데이터) + 기존 CSV 파일들을 읽어
Neo4j에 노드와 관계를 생성하는 스크립트 v2.

노드 종류:
  Building, Facility, Department, Transportation, Room (신규 마스터 데이터)
  Floor, Store (기존 파싱 데이터)

관계 종류:
  LOCATED_IN, HAS_FLOOR, HAS_ROOM, HAS_STORE (기존)
"""

import csv, os, re, time
import openpyxl
from neo4j import GraphDatabase

# ─── 접속 정보 ─────────────────────────────────────────────────
NEO4J_URI      = os.getenv("NEO4J_URI",      "bolt://localhost:7687")
NEO4J_USER     = os.getenv("NEO4J_USER",     "neo4j")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD", "campus-ai-password-2024")

BASE    = os.path.dirname(os.path.abspath(__file__))
XLSX    = os.path.join(BASE, "jbnu_campus_graphdb_data.xlsx")
COORDINATES_XLSX = os.path.join(BASE, "building_coordinates_20260809.xlsx")
CSV_DIR = BASE


# ─── 유틸 ─────────────────────────────────────────────────────
def s(v):
    """None → '' / 나머지 strip"""
    return str(v).strip() if v is not None else ""


def read_xlsx_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [s(h) for h in rows[0]]
    result = []
    for row in rows[1:]:
        record = {headers[i]: s(row[i]) for i in range(len(headers))}
        result.append(record)
    return result


def read_csv(filename):
    path = os.path.join(CSV_DIR, filename)
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def read_building_coordinates():
    """좌표 통합문서에서 건물번호별 단일 위도·경도를 추출합니다."""
    wb = openpyxl.load_workbook(COORDINATES_XLSX, read_only=True, data_only=True)
    points = {}

    def add(map_num, lat, lng, source_title):
        if not map_num or not isinstance(lat, (int, float)) or not isinstance(lng, (int, float)):
            return
        point = (float(lat), float(lng))
        existing = points.get(str(map_num))
        if existing and (existing["latitude"], existing["longitude"]) != point:
            raise ValueError(f"건물번호 {map_num}에 서로 다른 좌표가 있습니다.")
        points[str(map_num)] = {
            "map_num": str(map_num),
            "latitude": point[0],
            "longitude": point[1],
            "coordinate_source": source_title,
        }

    for row in wb["학교"].iter_rows(values_only=True):
        if len(row) < 9 or not row[2]:
            continue
        match = re.search(r"(?:^|\s)(\d+-\d+)(?:\s|$)", str(row[2]))
        if match:
            add(match.group(1), row[7], row[8], "학교")

    for row in wb["좌표 미매칭 목록"].iter_rows(values_only=True):
        if len(row) >= 8:
            add(row[0], row[3], row[4], "좌표 미매칭 목록")

    return list(points.values())


def run(driver, query, batch=None, params=None):
    with driver.session() as session:
        if batch is not None:
            session.run(query, batch=batch)
        else:
            session.run(query, **(params or {}))


# ══════════════════════════════════════════════════════════════
def load_all(driver):

    wb = openpyxl.load_workbook(XLSX)

    # 0. 초기화
    print("⚠️  기존 데이터 전체 초기화 중...")
    run(driver, "MATCH (n) DETACH DELETE n")
    print("  → 완료\n")

    # 인덱스
    print("📌 인덱스 생성 중...")
    for q in [
        "CREATE INDEX IF NOT EXISTS FOR (n:Building)       ON (n.building_id)",
        "CREATE INDEX IF NOT EXISTS FOR (n:Building)       ON (n.name)",
        "CREATE INDEX IF NOT EXISTS FOR (n:Facility)       ON (n.facility_id)",
        "CREATE INDEX IF NOT EXISTS FOR (n:Department)     ON (n.department_id)",
        "CREATE INDEX IF NOT EXISTS FOR (n:Transportation) ON (n.transportation_id)",
        "CREATE INDEX IF NOT EXISTS FOR (n:Room)           ON (n.room_id)",
        "CREATE INDEX IF NOT EXISTS FOR (n:Floor)          ON (n.building_name)",
        "CREATE INDEX IF NOT EXISTS FOR (n:Store)          ON (n.name)",
    ]:
        run(driver, q)
    print("  → 완료\n")

    # ── 1. Building ────────────────────────────────────────────
    data = read_xlsx_sheet(wb, "Buildings")
    print(f"🏛️  Building 노드 생성 중... ({len(data)}건)")
    batch = [{
        "id":            r["building_id:ID(Building)"],
        "name":          r["name"],
        "alias":         r["alias"],
        "zone":          r["zone"],
        "map_num":       r["map_location_number"],
        "campus":        r["campus"],
        "address":       r["address"],
        "phone":         r["phone"],
        "is_24h":        r["is_24_hours:boolean"],
        "floors":        r["floors"],
        "operating_hrs": r["operating_hours"],
        "source_url":    r["source_url"],
    } for r in data if r["building_id:ID(Building)"]]
    run(driver, """
        UNWIND $batch AS r
        MERGE (b:Building {building_id: r.id})
        SET b.name          = r.name,
            b.alias         = r.alias,
            b.zone          = r.zone,
            b.map_num       = r.map_num,
            b.campus        = r.campus,
            b.address       = r.address,
            b.phone         = r.phone,
            b.is_24h        = r.is_24h,
            b.floors        = r.floors,
            b.operating_hrs = r.operating_hrs,
            b.source_url    = r.source_url
    """, batch=batch)
    print(f"  → 완료\n")

    # ── 2. Facility ────────────────────────────────────────────
    data = read_xlsx_sheet(wb, "Facilities")
    print(f"🏥  Facility 노드 생성 중... ({len(data)}건)")
    batch = [{
        "id":          r["facility_id:ID(Facility)"],
        "type":        r["type"],
        "name":        r["name"],
        "floor":       r["floor"],
        "room_number": r["room_number"],
        "hours_wd":    r["operating_hours_weekday"],
        "hours_we":    r["operating_hours_weekend"],
        "hours_vac":   r["operating_hours_vacation"],
        "contact":     r["contact"],
        "description": r["description"],
        "source_url":  r["source_url"],
    } for r in data if r["facility_id:ID(Facility)"]]
    run(driver, """
        UNWIND $batch AS r
        MERGE (f:Facility {facility_id: r.id})
        SET f.type        = r.type,
            f.name        = r.name,
            f.floor       = r.floor,
            f.room_number = r.room_number,
            f.hours_wd    = r.hours_wd,
            f.hours_we    = r.hours_we,
            f.hours_vac   = r.hours_vac,
            f.contact     = r.contact,
            f.description = r.description,
            f.source_url  = r.source_url
    """, batch=batch)
    print(f"  → 완료\n")

    # ── 3. Department ──────────────────────────────────────────
    data = read_xlsx_sheet(wb, "Departments")
    print(f"🏫  Department 노드 생성 중... ({len(data)}건)")
    batch = [{
        "id":          r["department_id:ID(Department)"],
        "type":        r["type"],
        "name":        r["name"],
        "hours":       r["operating_hours"],
        "phone":       r["phone"],
        "email":       r["email"],
        "website":     r["website"],
        "description": r["description"],
        "source_url":  r["source_url"],
    } for r in data if r["department_id:ID(Department)"]]
    run(driver, """
        UNWIND $batch AS r
        MERGE (d:Department {department_id: r.id})
        SET d.type        = r.type,
            d.name        = r.name,
            d.hours       = r.hours,
            d.phone       = r.phone,
            d.email       = r.email,
            d.website     = r.website,
            d.description = r.description,
            d.source_url  = r.source_url
    """, batch=batch)
    print(f"  → 완료\n")

    # ── 4. Transportation ──────────────────────────────────────
    data = read_xlsx_sheet(wb, "Transportation")
    print(f"🚌  Transportation 노드 생성 중... ({len(data)}건)")
    batch = [{
        "id":          r["transportation_id:ID(Transportation)"],
        "type":        r["type"],
        "name":        r["name"],
        "routes":      r["routes:string[]"],
        "hours":       r["operating_hours"],
        "description": r["description"],
        "source_url":  r["source_url"],
    } for r in data if r["transportation_id:ID(Transportation)"]]
    run(driver, """
        UNWIND $batch AS r
        MERGE (t:Transportation {transportation_id: r.id})
        SET t.type        = r.type,
            t.name        = r.name,
            t.routes      = r.routes,
            t.hours       = r.hours,
            t.description = r.description,
            t.source_url  = r.source_url
    """, batch=batch)
    print(f"  → 완료\n")

    # ── 5. Room (신규 마스터) ──────────────────────────────────
    data = read_xlsx_sheet(wb, "Rooms")
    print(f"🚪  Room 노드 생성 중... ({len(data)}건)")
    batch = [{
        "id":          r["room_id:ID(Room)"],
        "room_number": r["room_number"],
        "name":        r["name"],
        "type":        r["type"],
        "floor":       r["floor"],
        "description": r["description"],
        "source_url":  r["source_url"],
    } for r in data if r["room_id:ID(Room)"]]
    run(driver, """
        UNWIND $batch AS r
        MERGE (rm:Room {room_id: r.id})
        SET rm.room_number = r.room_number,
            rm.name        = r.name,
            rm.type        = r.type,
            rm.floor       = r.floor,
            rm.description = r.description,
            rm.source_url  = r.source_url
    """, batch=batch)
    print(f"  → 완료\n")

    # ── 6a. 기존 파싱 Building (Floor 연결용) ─────────────────
    # Floor의 building_name이 기존 파싱 형식("단과대 · 건물명")이므로
    # 신규 xlsx Building과 별도로 파싱 Building도 함께 적재
    data = read_csv("nodes_building.csv")
    print(f"🏛️  Building(파싱) 노드 보완 중... ({len(data)}건)")
    run(driver, """
        UNWIND $batch AS r
        MERGE (b:Building {name: r.name})
        ON CREATE SET b.code          = r.code,
                      b.main_function = r.main_function,
                      b.source        = 'parsed'
    """, batch=data)
    print(f"  → 완료\n")

    # ── 6c. 건물 위도·경도 ───────────────────────────────────
    data = read_building_coordinates()
    print(f"📍 Building 좌표 갱신 중... ({len(data)}건)")
    run(driver, """
        UNWIND $batch AS r
        MATCH (b:Building {map_num: r.map_num})
        SET b.latitude          = r.latitude,
            b.longitude         = r.longitude,
            b.coordinate_source = r.coordinate_source
    """, batch=data)
    print(f"  → 완료\n")

    # ── 6b. 기존 파싱 데이터: Floor ───────────────────────────
    data = read_csv("nodes_floor.csv")
    print(f"🏢  Floor 노드 생성 중... ({len(data)}건)")
    run(driver, """
        UNWIND $batch AS r
        MERGE (f:Floor {building_name: r.building_name, floor: r.floor})
        SET f.building_code = r.building_code
    """, batch=data)
    print(f"  → 완료\n")

    # ── 7. 기존 파싱 데이터: Room (상세 호실) ─────────────────
    data = read_csv("nodes_room.csv")
    print(f"🚪  Room(상세 호실) 노드 생성 중... ({len(data)}건)")
    run(driver, """
        UNWIND $batch AS r
        MERGE (rm:Room {building_name: r.building_name, floor: r.floor, room_no: r.room_no})
        SET rm.name = r.name,
            rm.type = r.type
    """, batch=data)
    print(f"  → 완료\n")

    # ── 8. 기존 파싱 데이터: Store ─────────────────────────────
    data = read_csv("nodes_store.csv")
    print(f"🏪  Store 노드 생성 중... ({len(data)}건)")
    run(driver, """
        UNWIND $batch AS r
        MERGE (s:Store {name: r.name})
        SET s.type        = r.type,
            s.location    = r.location,
            s.hours       = r.hours,
            s.restriction = r.restriction,
            s.phone       = r.phone,
            s.note        = r.note
    """, batch=data)
    print(f"  → 완료\n")

    # ══ 관계 생성 ══════════════════════════════════════════════

    # ── 9. LOCATED_IN (Facility → Building) ───────────────────
    data = read_xlsx_sheet(wb, "Relationships")
    print(f"🔗  관계 생성 중... ({len(data)}건)")
    rels = [r for r in data if r[":TYPE"] == "LOCATED_IN"]
    run(driver, """
        UNWIND $batch AS r
        MATCH (src {facility_id: r.start})
        MATCH (tgt:Building {building_id: r.end})
        MERGE (src)-[:LOCATED_IN {description: r.desc}]->(tgt)
    """, batch=[{
        "start": r[":START_ID"],
        "end":   r[":END_ID"],
        "desc":  r["description"],
    } for r in rels])
    print(f"  → LOCATED_IN {len(rels)}개 완료\n")

    # ── 10. HAS_FLOOR (Building → Floor) ──────────────────────
    data = read_csv("rels_building_floor.csv")
    run(driver, """
        UNWIND $batch AS r
        MATCH (b:Building {name: r.building_name})
        MATCH (f:Floor {building_name: r.building_name, floor: r.floor})
        MERGE (b)-[:HAS_FLOOR]->(f)
    """, batch=data)
    print(f"  → HAS_FLOOR {len(data)}개 완료\n")

    # ── 11. HAS_ROOM (Floor → Room) ───────────────────────────
    data = read_csv("rels_floor_room.csv")
    run(driver, """
        UNWIND $batch AS r
        MATCH (f:Floor {building_name: r.building_name, floor: r.floor})
        MATCH (rm:Room {building_name: r.building_name, floor: r.floor, room_no: r.room_no})
        MERGE (f)-[:HAS_ROOM]->(rm)
    """, batch=data)
    print(f"  → HAS_ROOM {len(data)}개 완료\n")

    # ── 12. HAS_STORE (Building → Store) ──────────────────────
    run(driver, """
        MATCH (s:Store)
        MATCH (b:Building)
        WHERE s.location CONTAINS b.name OR b.name CONTAINS s.location
        MERGE (b)-[:HAS_STORE]->(s)
    """)
    run(driver, """
        MATCH (s:Store)
        WHERE NOT (s)<-[:HAS_STORE]-() AND s.location <> ''
        MERGE (b:Building {name: s.location})
        MERGE (b)-[:HAS_STORE]->(s)
    """)
    print(f"  → HAS_STORE 완료\n")


# ══════════════════════════════════════════════════════════════
def verify(driver):
    print("=" * 55)
    print("📊 최종 적재 결과 검증")
    print("=" * 55)
    queries = [
        ("Building",       "MATCH (n:Building)       RETURN count(n) AS c"),
        ("Facility",       "MATCH (n:Facility)       RETURN count(n) AS c"),
        ("Department",     "MATCH (n:Department)     RETURN count(n) AS c"),
        ("Transportation", "MATCH (n:Transportation) RETURN count(n) AS c"),
        ("Room",           "MATCH (n:Room)           RETURN count(n) AS c"),
        ("Floor",          "MATCH (n:Floor)          RETURN count(n) AS c"),
        ("Store",          "MATCH (n:Store)          RETURN count(n) AS c"),
        ("LOCATED_IN",     "MATCH ()-[r:LOCATED_IN]->() RETURN count(r) AS c"),
        ("HAS_FLOOR",      "MATCH ()-[r:HAS_FLOOR]->()  RETURN count(r) AS c"),
        ("HAS_ROOM",       "MATCH ()-[r:HAS_ROOM]->()   RETURN count(r) AS c"),
        ("HAS_STORE",      "MATCH ()-[r:HAS_STORE]->()  RETURN count(r) AS c"),
        ("Building+좌표",  "MATCH (n:Building) WHERE n.latitude IS NOT NULL AND n.longitude IS NOT NULL RETURN count(n) AS c"),
    ]
    with driver.session() as session:
        for label, q in queries:
            cnt = session.run(q).single()["c"]
            icon = "✅" if cnt > 0 else "❌"
            print(f"  {icon} {label:20s}: {cnt}개")


if __name__ == "__main__":
    print(f"🔌 Neo4j 연결 중... ({NEO4J_URI})")
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASSWORD))
    for attempt in range(5):
        try:
            driver.verify_connectivity()
            print("✅ 연결 성공!\n")
            break
        except Exception as e:
            print(f"  재시도 ({attempt+1}/5): {e}")
            time.sleep(3)
    else:
        print("❌ 연결 실패")
        exit(1)

    load_all(driver)
    verify(driver)
    driver.close()
    print("\n🎉 모든 작업 완료!")
